# USER GUIDE 📄 ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Read the README.md file for more context on this project.
# To run this file, install dependencies and run "python script.py" in your command line.
# The main part of this code you can change to try this yourself are those in the "⬇️ CONFIGURATION - change me! ⬇️" section.
# The DOM structure updates daily, so you may also need to check the current DOM structure being used and update the "todays_parent_element_type and todays_parent_element_class" variables
# To edit the URLs navigate to the papers.xlsx file.


# IMPORTS ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
import requests # HTTP requests
from bs4 import BeautifulSoup # Parsing HTML
import pandas as pd # Import and export Excel files
from openpyxl import load_workbook # Load Excel files for conditional colour formatting
from openpyxl.styles import PatternFill, Font # Pre-defined styles for conditional colour formatting



# ⬇️ CONFIGURATION - change me! ⬇️ ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# keywords
keywords = ["climate", "environment", "sustainability", "anxiety", "sensitivity", "mitigation"]
keyword_results = {keyword: [] for keyword in keywords}  # Initialize a dictionary for keyword results

# request information
timeout_seconds = 10 # timeout for the HTTP requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

# abstract character lengths which trigger warning flags
min_abstract_char_length = 450
max_abstract_char_length = 2500

# file names/paths
output_file_name = 'output.xlsx'
input_file_name = 'papers.xlsx'

# colours
green_font = Font(color="55C233")
red_font = Font(color="FF0000")

# ⬇️ change the DOM element structure below⬇️
# todays_parent_element_type = "div" 
todays_parent_element_type = "section"
# todays_parent_element_class = "tsec"
todays_parent_element_class = "abstract"




# SCRIPT RUNNING MESSAGE ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
print("Running script.py\n")


# IMPORT SPREADSHEET DATA ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
data_frame = pd.read_excel(input_file_name, sheet_name='urls')  # You can specify the sheet name or index

urls = []
if 'URLS' in data_frame.columns:
    for url in data_frame['URLS']:
        urls.append(url)
else:
    print("The 'URLS' column does not exist in the DataFrame.")



# ... KEYWORD AND ABSTRACT PLACEHOLDERS ///////////////
abstract_data = {
    "abstract_text": [],
    "abstract_length": [],
    "abstract_concerns": []
}


# LOOPING THROUGH EACH URL REQUEST ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
for index, url in enumerate(urls):
    try:
        # GENERIC REQUEST AND PASSING///////////////////////////////////////
        print(f"\n_________________________ FETCHING URL ({index}) - {url} ______________________________________________\n")
        page_to_scrape = requests.get(url, headers=HEADERS, timeout=timeout_seconds)
        page_to_scrape.raise_for_status()  # Raises an HTTPError for bad responses (4xx, 5xx)
        soup = BeautifulSoup(page_to_scrape.content, 'html.parser')


  
        # DATA HANDLING FOR ARTICLES ON ---- NIH ----- WEBSITE /////////////////////////////////////////////////////////////////////////////////////////////////////
        # NIH STRATEGY 1: find div that has class "tsec" and a child h2 element with text "Abstract"
        # NIH STRATEGY 2: on some days, the structure changes to use sections with class "abstract" and the same h2 element

        all_divs = soup.findAll(todays_parent_element_type)
        abstract_found = False # flag
        # step: loop through all divs
        for div in all_divs:
            # step: check if todays_parent_element_type has class todays_parent_element_class
            if ('class' in div.attrs):
                div_classes = div.attrs['class']
                if todays_parent_element_class in div_classes:
                    # step: get all direct h2 children
                    all_h2s = div.findAll("h2", recursive=False) # recursive = False allows finding direct parents only, preventing duplicate prints
                    # step: Find the first h2 with 'Abstract' or 'Summary'
                    first_h2 = next((h2 for h2 in all_h2s if 'Abstract' in h2.getText() or 'Summary' in h2.getText()), None) # A shorthand form of looping through the h2s

                    if first_h2:
                        abstract_found = True
                        div_text = div.getText().replace("Abstract", "").replace("Summary", "").strip().lower()
                        # record abstract scrape and length analysis information in variables
                        abstract_data["abstract_text"].append(div_text)
                        abstract_data["abstract_length"].append(len(div_text))

                        if len(div_text) < min_abstract_char_length:
                            abstract_data["abstract_concerns"].append("Short - check manually")
                        elif len(div_text) > max_abstract_char_length:
                            abstract_data["abstract_concerns"].append("Long - check manually")
                        else:
                            abstract_data["abstract_concerns"].append("Length: OK")

                        # record keyword check information in variables
                        for keyword in keywords:
                            if keyword in div_text:
                                print(f"Keyword '{keyword}' found ✅")
                                keyword_results[keyword].append("present")
                            else:
                                print(f"Keyword '{keyword}' NOT found ❌")
                                keyword_results[keyword].append("missing")

                        
                        break  # Exit the loop after processing the first instance (to prevent analysing Summaries provided at the end of papers and analysing two text-sources for some papers)


        if(abstract_found == False):
            print("!!! Cannot find an Abstract, Summary or the parent that contains them.")
            # record error messages in variables for the output file
            for keyword in keywords:
                keyword_results[keyword].append("ERROR - ABSTRACT NOT FOUND")
            for abstract in abstract_data:
                abstract_data[abstract].append("ERROR - ABSTRACT NOT FOUND")


        # END OF DATA HANDLING FOR ARTICLES ON NATIONAL LIBRARY OF MEDICINE WEBSITE /////////////////////////////////////////////////////////////////////////////////////////////////////
    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except requests.exceptions.ConnectionError as conn_err:
        print(f"Connection error occurred: {conn_err}")
    except requests.exceptions.Timeout as timeout_err:
        print(f"Timeout error occurred: {timeout_err}")
    except requests.exceptions.RequestException as req_err:
        print(f"Error occurred: {req_err}")



# Exporting analysis ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
try: 
    # Update DataFrame with abstract and keyword data //////////////////////////////////////////////////////////
    for abstract in abstract_data:
        data_frame[abstract] = abstract_data[abstract]
    for keyword in keywords:
        data_frame[keyword] = keyword_results[keyword]

    # Export to Excel ////////////////////////////////////////////////////////////
    data_frame.to_excel(output_file_name, index=False, sheet_name='urls')  # index=False to omit row indices
    print("\n")

    print("\n Attempting to export...")


    print("SUCCESS! Export to Excel successful.")

except Exception as e:
    print("ERROR! problem exporting analysis - do you have the file open?\n", e)





# Apply conditional formatting ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
try:
    workbook = load_workbook(output_file_name)
    sheet = workbook.active

    # Apply styles based on the values ////////////////////////////////////////////////////////
    for row in range(2, len(data_frame) + 2):  # Start from the second row to skip the header
        # Check abstract concerns ////////////////////////////////////////////////////////
        abstract_concern_cell = sheet[f"D{row}"]  # Assuming abstract concerns are in column D
        if "OK" in str(abstract_concern_cell.value):
            abstract_concern_cell.font = green_font
        elif str(abstract_concern_cell.value.startswith("Short")) or abstract_concern_cell.value.startswith("Long"):
            abstract_concern_cell.font = red_font

        # Check keywords ////////////////////////////////////////////////////////
        for col in range(4, len(keywords) + 5):  # Assuming keywords start from column D, 4 represents colomn D
            keyword_cell = sheet.cell(row=row, column=col)
            if keyword_cell.value == "present":
                keyword_cell.font = green_font
            elif keyword_cell.value == "missing":
                keyword_cell.font = red_font

    # Save the workbook ////////////////////////////////////////////////////////
    workbook.save(output_file_name)
    print("\n")

    print("\n Attempting to apply conditional formatting...")

    print("SUCCESS! Colour code successful.\n")

except Exception as e:
    print("ERROR! problem colour coding the analysis\n", e)